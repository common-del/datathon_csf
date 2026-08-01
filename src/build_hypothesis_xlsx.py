"""Build outputs/HYPOTHESIS_REGISTER.xlsx: all 36 hypotheses, methodology, finding, verdict, assumptions.
Numbers are read from outputs/tables/hypothesis_menu.csv so nothing is retyped by hand.
Run from repo root: python src/build_hypothesis_xlsx.py
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__)); ROOT = os.path.abspath(os.path.join(HERE, ".."))
os.chdir(ROOT)
MENU = pd.read_csv("outputs/tables/hypothesis_menu.csv").set_index("id")

# Human layer: theme, plain question, method, assumptions, kill-switch, where used.
# Effect and caveat text come from the pipeline CSV, never retyped.
M = {
"EH1": ("1. Is the grade 6 decline real?",
  "Does the grade 6 collapse survive if we only look at places that took part every year?",
  "Kept only the Gram Panchayats that fielded at least 10 grade 6 children in all three years (2,182 GPs). Compared their 2022-23 and 2024-25 competency scores, student-weighted. If new weaker places caused the fall, it should shrink or vanish here.",
  "Panel GPs are early joiners, so better organised on average. That bias would flatter the trend, not worsen it. Items change yearly, so comparison is at competency level.",
  "If the 2,182 panel GPs were unrepresentative in a way that inflates decline, e.g. all from one collapsing region. Check: rerun by revenue division.",
  "Report 4.0 (lead finding) | Slide 4 | Policy note para 1"),
"EH2": ("1. Is the grade 6 decline real?",
  "Does the one piece of good news, grade 4 division improving, also survive?",
  "Identical constant-GP panel construction as EH1, applied to grade 4 division (2,278 GPs).",
  "Same panel caveats as EH1. The raw gain and the panel result use the same competency mapping.",
  "Nothing visible. Two panels built identically gave opposite answers about their own headlines, which is the test working.",
  "Report 4.0 and section 5 | Slide 4 | Killed our own good news"),
"EH3": ("1. Is the grade 6 decline real?",
  "Does the decline show up in districts where coverage barely moved?",
  "Ranked districts by absolute change in coverage 2022-23 to 2024-25, took the bottom third (8 districts, mean change 3.9pp), and measured their grade 6 competency change.",
  "Unweighted district means. District-level coverage stability does not guarantee stability of reach within the district.",
  "Only 8 districts, so noisy. But EH1 and EH6 use different logic and land in the same range.",
  "Report 4.0 | Slide 4"),
"EH4": ("1. Is the grade 6 decline real?",
  "Do the districts that expanded coverage most show the biggest score falls?",
  "Correlated each district's change in coverage against its change in mean score, 2022-23 to 2024-25, rural State Government basis on both sides. This is the selection story's own preferred test.",
  "Ecological correlation at district level. Coverage growth also tracks administrative effort, which could push the other way.",
  "Only 25 districts, so a small effect could hide. But we are asking whether it explains a 16pp fall, and it does not.",
  "Report 4.0 and section 5 | Slide 4 | Kills the selection story"),
"EH5": ("1. Is the grade 6 decline real?",
  "If you statistically remove coverage change, how much decline is left?",
  "District-level OLS: regressed change in grade 6 multiplication on change in coverage, then read the intercept, i.e. the predicted decline for a district whose coverage did not move.",
  "Linear control only, at district level. Selection happening school by school inside a district is invisible to this test.",
  "A strongly non-linear coverage effect. EH1 handles the within-district case a different way.",
  "Report 4.0 | claims.json claim-eh5-intercept"),
"EH6": ("1. Is the grade 6 decline real?",
  "Does the decline appear in districts that were already deeply covered from the start?",
  "Selected the 11 districts already at 40% coverage or higher in 2022-23. These had least room for composition to dilute the average, since most eligible children were already in.",
  "Small n. Deep-start districts may differ in administrative capacity or geography.",
  "If deep-start districts share a confound that drives decline independently. Four designs now agree within 18-22pp.",
  "Report 4.0 | Slide 4"),
"EH7": ("4. Averages and league tables mislead",
  "Does the district league table change once you account for who was assessed?",
  "Regressed 2024-25 district mean score on district coverage, then re-ranked districts on the residual plus the state mean. Compared raw rank against adjusted rank.",
  "The adjustment is a descriptive straight-line correction, not a causal fix. Published as a companion table, never as the true ranking.",
  "If the coverage-score relationship is non-linear, the specific rank shifts change. The general point, that ranks are coverage-sensitive, would survive.",
  "Report 4.7 | Recommendation 3 | outputs/tables/league_coverage_adjusted.csv"),
"EH8": ("2. Is the girls' lead real?",
  "Where more girls are tested than boys, is the girls' measured lead bigger? (It is smaller.)",
  "For each district-year, computed the girls-minus-boys gap in assessment coverage (verified UDISE rural State Govt denominator) and the girls-minus-boys gap in score, then correlated the two across 80 district-years.",
  "District-level; selection within a school is unobserved. We first reported this with the sign flipped, because a pivot sorts to boys, girls while the score gap is girls-minus-boys. Caught in review and fixed in src/extra_hypotheses.py.",
  "District-level only. But the corrected sign is what makes EH8 and EH9 agree: if over-assessment shrank the measured gap, balanced districts should show a bigger one, and they do (+3.6 vs +2.8).",
  "Report 4.5 | Slide 7 | Strengthens the gender finding rather than weakening it"),
"EH9": ("2. Is the girls' lead real?",
  "Do girls still lead where boys and girls are tested at equal rates?",
  "Restricted to the 25 district-years where the boy-girl coverage gap is within 2pp, then recomputed the score gap. If the lead is a participation mirage it should shrink here.",
  "Balance is defined at district level on the verified UDISE denominator. It does not guarantee balance in every school. Sample shrinks under the filter.",
  "If the 25 balanced district-years are systematically high-performing. Check: compare their mean score against the state mean.",
  "Report 4.5 | Slide 7 | claims.json claim-eh9-genderbal"),
"EH10": ("2. Is the girls' lead real?",
  "Is the girls' lead steady over time while coverage doubled?",
  "Computed the mean district-level girls-minus-boys gap separately for each of the three years. A participation artefact should wobble as participation changes.",
  "Unweighted mean of district gaps. Items change yearly.",
  "A coincidence where two opposing changes cancel out across three years. Unlikely.",
  "Report 4.5 | Slide 7"),
"EH11": ("3. Can we trust the instrument?",
  "Does our district map agree with ASER 2024, rural against rural?",
  "Spearman rank correlation between our 2024-25 district mean and ASER 2024 district arithmetic estimates. Both universes are rural, so this is like-for-like, not an analogy. Tested against both the Std 3-5 subtraction and Std 6-8 division scales.",
  "ASER district samples are roughly 600 children each, so both rankings are noisy. Our 2024-25 round excludes 4 districts.",
  "If the agreement were driven by a handful of extreme districts. Check: drop top and bottom 3 and rerun.",
  "Report 4.8 | Slide 8 | Validation of the whole submission"),
"EH12": ("3. Can we trust the instrument?",
  "Is our agreement with ASER better in well-covered districts?",
  "Computed absolute rank disagreement against ASER for each district, then correlated it with that district's coverage. Our theory was that thin coverage should cause disagreement.",
  "Comparing two noisy rankings, so the disagreement measure is itself noisy.",
  "We were wrong. r = -0.01. Reported as a discard because our hypothesis failed, not because the data is bad.",
  "Report section 5 | Our own theory, disproved"),
"EH13": ("5. Where in maths do children fall off?",
  "Which single competency tells you the most about a child's overall performance?",
  "For each competency, correlated it against the mean of the OTHER competencies, per grade-year, then averaged across the 9 grade-years. Excluding the competency from its own comparison removes the part-whole overlap that would rig the result.",
  "Correlation, not sequencing. The top three (division 0.67, measurement 0.67, multiplication 0.66) are within 0.01, so naming a single winner overstates the gap.",
  "Nothing breaks it, but the honest phrasing is that the multiplicative family carries most information, not that division specifically is best.",
  "Report 4.2 | Supports the short-screener recommendation"),
"EH14": ("5. Where in maths do children fall off?",
  "Is there a prerequisite ladder in maths mastery?",
  "For each pair of competencies, computed the share of children mastering B among those who mastered A, against the share among those who did not. The difference is the lift.",
  "Conditional probability, not proven causal ordering. Both could reflect general ability.",
  "General ability driving both. But the ordering matches the curriculum's own logic, so it still guides what to teach first.",
  "Report 4.2 | Slide 5 | Recommendation 2 (sequencing)"),
"EH15": ("4. Averages and league tables mislead",
  "Are the weakest places also the most unequal inside themselves?",
  "Correlated GP mean score against GP standard deviation across 5,599 GPs with at least 15 students, all years pooled.",
  "Scores near the ceiling mechanically compress spread, which inflates this negative correlation. Direction is trustworthy, magnitude is not.",
  "The ceiling effect. Do not quote -0.62 as clean. A cleaner test would use a variance-stabilising transform.",
  "Report 4.4 | Argues for whole-class support in weak GPs"),
"EH16": ("4. Averages and league tables mislead",
  "Does the measured floor fall where coverage grew?",
  "Correlated district change in coverage against district change in the 10th-percentile score, 2022-23 to 2024-25.",
  "Both floor and coverage move with district administrative effort, so this is association, not attribution.",
  "If floor change were driven by something else correlated with expansion. EH24 tests the same idea a second way and agrees.",
  "Report 4.4 | Recommendation 3"),
"EH17": ("6. Geography, history and money",
  "Did districts hit by the 2023 drought decline more the following year?",
  "Attempted to split districts into drought-declared and not, using the verified events timeline, then compare 2023-24 to 2024-25 change.",
  "The declaration covered 223 taluks, which is most of Karnataka. Our data joins at district, the shock is at taluk level.",
  "Untestable as designed. A taluk-level assessment file would make it testable.",
  "Report section 5 and limitations | Reported as untestable, not as null"),
"EH18": ("6. Geography, history and money",
  "Do the over-performing blocks cluster, or are they scattered luck?",
  "Flagged 22 blocks whose actual score beats the structural prediction by a wide margin, then ran a chi-square test on their distribution across districts against expected counts proportional to each district's block count.",
  "The block prediction model explains only 26% of variation (cv R2 = 0.26), so the residuals used to flag bright spots are noisy.",
  "Model misspecification could cluster residuals by district if a district-level covariate is missing from the model.",
  "Report 4.9 | Slide 8 | Recommendation 2 (demonstration sites)"),
"EH19": ("6. Geography, history and money",
  "Is the Kalyana Karnataka gap closing or widening?",
  "Computed the student-weighted mean score for the 7 Article 371J districts against all other districts, separately for each of the three years.",
  "Coverage grew fastest in some KK districts, so part of the widening may be deeper reach rather than falling learning.",
  "The coverage confound. State it alongside, as the report does. A constant-GP version of this test would settle it.",
  "Report 4.3 | Slide 10 | Recommendation 1"),
"EH20": ("3. Can we trust the instrument?",
  "Are any of the test questions bad?",
  "Computed point-biserial discrimination for all 180 item-year-grade combinations: does each question separate stronger children from weaker ones?",
  "Low discrimination can also mean a question everyone got right or everyone got wrong, so the threshold is a screen, not a verdict.",
  "Nothing. Zero items fall below 0.15. Reported as a positive: the instrument is clean.",
  "Report 4.8 and section 5 | Supports instrument credibility"),
"EH21": ("2. Is the girls' lead real?",
  "Do any individual questions behave differently for boys and girls?",
  "Screened all 180 item-year-grades for a raw girls-minus-boys gap of 5pp or more.",
  "Crude screen. Proper differential item functioning requires matching boys and girls on total score first.",
  "The screen confuses genuine item bias with real ability differences. That is why we call it weak and only flag items for review.",
  "Report section 5 | Flagged for Akshara's assessment team"),
"EH22": ("1. Is the grade 6 decline real?",
  "Do the newly reached places actually have weaker children?",
  "Identified the 655 GPs that first appear in 2024-25, compared their grade 6 mean against veteran GPs. This tests the unstated assumption the entire selection story rests on.",
  "New GPs may be smaller or may have sent only their strongest children on a first outing. Either way it is not evidence for the selection story.",
  "First-outing selection within new GPs. Even so, the selection story requires new places to score lower, and they score higher.",
  "Report 4.0 and section 5 | Selection story runs backwards"),
"EH23": ("4. Averages and league tables mislead",
  "Are the districts at the top of the table actually the best?",
  "Ranked all 29 districts on pooled all-year student-weighted mean, then placed each district's total coverage next to its rank. Also checked which districts are missing from the 2024-25 round.",
  "Thin samples are noisy in both directions. We are not claiming these districts are secretly weak, only that their rank is not a district fact.",
  "Nothing breaks the arithmetic. The interpretation would change if the thin samples were random rather than self-selected, which is implausible for a voluntary contest.",
  "Report 4.7 | Strongest single argument for Recommendation 3"),
"EH24": ("4. Averages and league tables mislead",
  "Does mean-versus-floor divergence concentrate where coverage grew fastest?",
  "Split districts into three equal groups by coverage growth, then compared the average gap between mean movement and 10th-percentile movement in each group.",
  "District-level, same selection caveat as EH16.",
  "Tercile boundaries are arbitrary. The monotonic pattern across all three groups makes an artefact less likely.",
  "Report 4.4 | Second test of the same idea as EH16"),
"H1": ("6. Geography, history and money",
  "Do blocks with more crowded classrooms score worse?",
  "Correlated block-level government-school pupil-teacher ratio against block mean score, partially controlling for Census block literacy.",
  "PTR proxies remoteness and school size as well as teacher availability. Block join to UDISE is 68%.",
  "Direction is right and p = 0.07 is borderline. A larger block sample might clear the bar. We do not present it alone.",
  "Not presented alone | PTR enters the submission through H5"),
"H2": ("6. Geography, history and money",
  "Do non-Kannada-medium belts score worse on a Kannada-language test?",
  "Correlated the share of non-Kannada-medium government schools in a block against block mean score.",
  "We measured the medium of schools in the block, not the language spoken at home by the child who sat the test. That is a real gap between measure and concept.",
  "The measure may be too coarse. A home-language variable, or an item-level split of word problems against pure computation, would test it properly.",
  "Report 4.6 and section 5 | Reported as a null on a plausible equity worry"),
"H4": ("6. Geography, history and money",
  "Does early childhood stunting predict this cohort's maths?",
  "Correlated NFHS-5 district under-5 stunting against district mean score. The tested children were born 2012-2016, so they ARE the NFHS-5 under-5 cohort. The exposure timing lines up exactly.",
  "NFHS-5 is district-level and mixes urban with rural, while our universe is rural only. That mismatch may be enough to wash out a real effect.",
  "Aggregation. A rural-only, block-level nutrition measure could still find something. We report the null at the level we can test.",
  "Report section 5 | Our best-designed question, and it returned nothing"),
"H5": ("6. Geography, history and money",
  "Is the Kalyana Karnataka gap explained by measured inputs, or is it a mystery?",
  "Computed the raw score gap between the 7 Article 371J districts and the rest, then recomputed it after controlling for pupil-teacher ratio, adult literacy and under-5 stunting.",
  "Three controls only. A decade of 371J funds and the 2023-24 KKRDB education year push in the opposite direction within the window.",
  "Omitted variables. But the direction of the finding is hopeful: most of the gap is inputs a government can change.",
  "Report 4.3 | Slide 10 | Recommendation 1 | Must be shown next to H11"),
"H6": ("2. Is the girls' lead real?",
  "Does the gender gap change as the maths gets harder?",
  "Split competencies into easiest and hardest thirds by overall mastery, then compared the girls-minus-boys gap in each tier and rank-correlated gap against difficulty.",
  "Tier boundaries are a third-and-third split of 11 competencies, so each tier is small.",
  "Nothing. The gap is flat across difficulty (+2.3 vs +2.6). The stereotype is not in this data.",
  "Report 4.5 and section 5 | Kills a popular assumption"),
"H7": ("4. Averages and league tables mislead",
  "Do averages move while the weakest children stay put?",
  "Compared each district's change in mean against its change in 10th-percentile score, first year to last, and counted districts where the floor lagged the mean.",
  "Items changed yearly, so this is framed as shape and ranking shifts, not point gains.",
  "Read alone it looks like an equity failure. EH16 shows part of it is who newly arrived, so the two must be read together.",
  "Report 4.4 | Slide 6 | Recommendation 3"),
"H8": ("6. Geography, history and money",
  "Do children in teacher-starved blocks progress more slowly?",
  "Compared cohort progression (same place, same cohort, successive years) between the most and least crowded quartiles of blocks by PTR.",
  "Cohorts are synthetic: same place and cohort, different individual children. Composition can shift between years.",
  "Nothing. A difference of 0.04pp with p = 0.97 is as null as results get.",
  "Report section 5"),
"H9": ("1. Is the grade 6 decline real?",
  "Can we test who-got-tested bias by joining GP names to UDISE?",
  "Attempted a GP-name join to UDISE enrolment to compare tested against enrolled children at GP level. An automatic reliability gate refused the join.",
  "GP names repeat across districts and the fuzzy match reached only 37%, below the 60% usable threshold.",
  "Not fixable with available data. We answered the same question better at district level, in EH1 to EH6 and EH22.",
  "Report 6b | Shows exactly where our data runs out"),
"H10": ("6. Geography, history and money",
  "Do government schools score worse where private schools are common?",
  "Correlated the share of private schools in a block against block mean government-school score, testing whether private schools cream off advantaged children.",
  "This is a composition and selection story, not a measure of school quality. Block join is 68%.",
  "r = -0.14, p = 0.13. Direction is plausible but the evidence is not there at this sample size.",
  "Report section 5"),
"H11": ("6. Geography, history and money",
  "Does the learning map follow pre-1956 administrative borders?",
  "Grouped districts by the administration they belonged to before the States Reorganisation Act (Bombay, Hyderabad, Madras, Mysore, Coorg) and ran ANOVA on group means.",
  "We coded the legacy groupings ourselves from historical literature. Verify before quoting. Legacy correlates with present-day inputs, so this is not an independent explanation.",
  "Confounding with current inputs, which H5 shows explain most of the gap. Dangerous if presented without H5 beside it.",
  "Report 4.3 context | Never present alone"),
"H12": ("3. Can we trust the instrument?",
  "Does our district map replicate in PARAKH RS 2024?",
  "Spearman rank correlation between our district means and PARAKH RS 2024 grade 6 maths, restricted to the state-government subgroup so the management universe matches.",
  "PARAKH publishes whole integers on roughly 1,300 schools statewide, so districts tie frequently, which caps the achievable correlation.",
  "Ties and rounding. rho = 0.51 on a rounded scale is arguably stronger than it looks.",
  "Report 4.8 | Slide 8 | Second external validation"),
"H13": ("2. Is the girls' lead real?",
  "Do district-level gender gaps replicate in PARAKH RS 2024?",
  "Correlated our district girls-minus-boys gap against PARAKH's, and separately counted how often the two agree on direction.",
  "Both instruments are cross-sectional. PARAKH's district gender cells are small and rounded.",
  "District-level gender gaps are noisy in both datasets. We report the 76% direction agreement and the weak correlation together, and claim only the former.",
  "Report 4.8 | Reported with its own weakness attached"),
}

rows = []
for hid, (theme, question, method, assume, kill, used) in M.items():
    r = MENU.loc[hid]
    rows.append({"ID": hid, "Theme": theme, "Plain-English question": question,
                 "Formal hypothesis (as tested)": r["hypothesis"], "Methodology": method,
                 "Finding (numbers)": r["effect"], "Verdict": r["verdict"],
                 "Assumptions and caveats": (assume + " || Pipeline caveat: " + str(r["caveat"])),
                 "What would break this": kill,
                 "Policy meaning": r["policy_meaning"], "Where used": used,
                 "Source script": "src/extra_hypotheses.py" if hid.startswith("EH") else "src/day1_verdicts.py"})
df = pd.DataFrame(rows)
order = {"SUPPORTED": 0, "WEAK": 1, "DISCARD": 2}
df = df.sort_values(["Theme", "Verdict"], key=lambda s: s.map(order) if s.name == "Verdict" else s).reset_index(drop=True)

# ------------------------------------------------------------------ workbook
wb = Workbook()
ARIAL = "Arial"
INK = "FF123B47"; HEADFILL = PatternFill("solid", fgColor="FF123B47")
GREENF = PatternFill("solid", fgColor="FFD9F0D3"); AMBERF = PatternFill("solid", fgColor="FFFDF0D5")
REDF = PatternFill("solid", fgColor="FFFBDDD8"); BANDF = PatternFill("solid", fgColor="FFF4F8F8")
THIN = Side(style="thin", color="FFC9D4D6"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ws = wb.active; ws.title = "Hypotheses"
ws.append(list(df.columns))
for c in range(1, len(df.columns) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(name=ARIAL, bold=True, size=10, color="FFFFFFFF")
    cell.fill = HEADFILL
    cell.alignment = Alignment(wrap_text=True, vertical="center"); cell.border = BORDER
ws.row_dimensions[1].height = 34

VFILL = {"SUPPORTED": GREENF, "WEAK": AMBERF, "DISCARD": REDF}
for i, rec in enumerate(df.to_dict("records"), start=2):
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=i, column=j, value=rec[col])
        cell.font = Font(name=ARIAL, size=9, bold=(col in ("ID", "Verdict")))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BORDER
        if col == "Verdict":
            cell.fill = VFILL[rec["Verdict"]]; cell.alignment = Alignment(horizontal="center", vertical="center")
        elif i % 2 == 0:
            cell.fill = BANDF
    ws.row_dimensions[i].height = 112

W = {"ID": 7, "Theme": 22, "Plain-English question": 34, "Formal hypothesis (as tested)": 34,
     "Methodology": 52, "Finding (numbers)": 44, "Verdict": 12, "Assumptions and caveats": 50,
     "What would break this": 44, "Policy meaning": 42, "Where used": 28, "Source script": 20}
for j, col in enumerate(df.columns, start=1):
    ws.column_dimensions[get_column_letter(j)].width = W[col]
ws.freeze_panes = "C2"
ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(df.columns)), len(df) + 1)

# ------------------------------------------------------------------ summary (formulas, not hardcoded)
s = wb.create_sheet("Summary")
n = len(df) + 1
s["A1"] = "Hypothesis register: summary"; s["A1"].font = Font(name=ARIAL, bold=True, size=13, color=INK)
s["A2"] = "Counts recalculate from the Hypotheses tab. Source of every number: outputs/tables/hypothesis_menu.csv, produced by src/run_all.py then src/day1_verdicts.py then src/extra_hypotheses.py (seed 20260801)."
s["A2"].font = Font(name=ARIAL, size=9, italic=True); s["A2"].alignment = Alignment(wrap_text=True)
s.merge_cells("A2:D2"); s.row_dimensions[2].height = 30
hdr = ["Verdict", "Count", "Share of all", "What it means"]
for j, h in enumerate(hdr, start=1):
    c = s.cell(row=4, column=j, value=h)
    c.font = Font(name=ARIAL, bold=True, size=10, color="FFFFFFFF"); c.fill = HEADFILL; c.border = BORDER
meanings = ["Effect is sizeable, unlikely to be chance, and survived a test built to break it. Safe to present.",
            "Visible but small, or failed one check. Mention with the caveat; do not build a slide on it.",
            "No usable signal, or the test could not be run. Reported openly in report.pdf section 5."]
for i, (v, mng) in enumerate(zip(["SUPPORTED", "WEAK", "DISCARD"], meanings), start=5):
    s.cell(row=i, column=1, value=v).font = Font(name=ARIAL, bold=True, size=10)
    s.cell(row=i, column=1).fill = VFILL[v]
    s.cell(row=i, column=2, value="=COUNTIF(Hypotheses!$G$2:$G$%d,A%d)" % (n, i)).font = Font(name=ARIAL, size=10)
    s.cell(row=i, column=3, value="=B%d/$B$8" % i).font = Font(name=ARIAL, size=10)
    s.cell(row=i, column=3).number_format = "0.0%"
    s.cell(row=i, column=4, value=mng).font = Font(name=ARIAL, size=9)
    s.cell(row=i, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    s.row_dimensions[i].height = 40
    for j in range(1, 5): s.cell(row=i, column=j).border = BORDER
s.cell(row=8, column=1, value="TOTAL").font = Font(name=ARIAL, bold=True, size=10)
s.cell(row=8, column=2, value="=SUM(B5:B7)").font = Font(name=ARIAL, bold=True, size=10)
s.cell(row=8, column=3, value="=B8/$B$8").font = Font(name=ARIAL, bold=True, size=10)
s.cell(row=8, column=3).number_format = "0.0%"
for j in range(1, 5): s.cell(row=8, column=j).border = BORDER

s.cell(row=10, column=1, value="By theme").font = Font(name=ARIAL, bold=True, size=11, color=INK)
for j, h in enumerate(["Theme", "Total", "Supported", "Weak", "Discarded"], start=1):
    c = s.cell(row=11, column=j, value=h)
    c.font = Font(name=ARIAL, bold=True, size=10, color="FFFFFFFF"); c.fill = HEADFILL; c.border = BORDER
for i, th in enumerate(sorted(df["Theme"].unique()), start=12):
    s.cell(row=i, column=1, value=th).font = Font(name=ARIAL, size=9)
    s.cell(row=i, column=2, value='=COUNTIF(Hypotheses!$B$2:$B$%d,$A%d)' % (n, i)).font = Font(name=ARIAL, size=9)
    for k, v in enumerate(["SUPPORTED", "WEAK", "DISCARD"], start=3):
        s.cell(row=i, column=k, value='=COUNTIFS(Hypotheses!$B$2:$B$%d,$A%d,Hypotheses!$G$2:$G$%d,"%s")' % (n, i, n, v)).font = Font(name=ARIAL, size=9)
    for j in range(1, 6): s.cell(row=i, column=j).border = BORDER
for col, w in zip("ABCDE", [40, 10, 12, 10, 12]): s.column_dimensions[col].width = w

# ------------------------------------------------------------------ how to read
h = wb.create_sheet("How to read this")
notes = [
 ("What this file is", "Every hypothesis we tested on the Akshara GP Maths Contest data, whether it survived or not. 36 in total. Testing in public is deliberate: a question answered honestly and discarded is worth more than a question avoided."),
 ("Where the numbers come from", "The Finding, Verdict, formal hypothesis text and pipeline caveat are read directly from outputs/tables/hypothesis_menu.csv by src/build_hypothesis_xlsx.py. They are not retyped by hand. The plain-English question, methodology, assumptions and break-conditions are written by the team."),
 ("How to reproduce", "From the repo root: python src/run_all.py, then python src/day1_verdicts.py, then python src/extra_hypotheses.py. Fixed seed 20260801. Runs offline in about three minutes."),
 ("Verdict thresholds", "SUPPORTED needs a sizeable effect, statistical clarity, and survival of a robustness check. WEAK means small, or one check failed. DISCARD means no usable signal or the test could not be run. Thresholds are stated in the docstring of each source script so a judge can check them."),
 ("Causal language", "This is observational data. Every judgment in the report is phrased as consistent-with. Nothing here proves causation, and nothing claims to."),
 ("The cross-year caveat", "The 20 questions change every year within a constant competency framework. All cross-year statements are made at competency or ranking level, never as raw item-score gains, and the headline decline is corroborated by two independent external assessments (ASER 2024 and PARAKH RS 2024)."),
 ("The universe", "Rural State Government schools only, grades 4 to 6, 1,379,087 student records across three years. Every UDISE denominator filters to management_group = State Government and rural_urban = Rural. This makes ASER, which is also rural-only, a like-for-like benchmark."),
 ("Known weaknesses we would raise ourselves", "EH13 names division as the best single proxy at 0.67, but measurement is also 0.67 and multiplication 0.66; the honest phrasing is that the multiplicative family carries most information. EH15's correlation of -0.62 is inflated by a ceiling effect; the direction holds, the magnitude does not."),
 ("Two errors we found and fixed in review", "First, src/coverage.py joined from the assessed side, silently dropping every district-grade-year that enrolled children but assessed none. That inflated reported coverage from 37.8% to 41.8% and the yearly series from 25.1/39.0/49.9 to 26.8/43.8/56.2. Fixed; the corrected denominator now reconciles to the independently cross-validated UDISE gender file to the child. Second, EH8 carried a sign flip. Both are documented here rather than quietly corrected, and neither changed a single verdict on the coverage-robustness tests."),
]
h["A1"] = "How to read this register"; h["A1"].font = Font(name=ARIAL, bold=True, size=13, color=INK)
for i, (k, v) in enumerate(notes, start=3):
    h.cell(row=i, column=1, value=k).font = Font(name=ARIAL, bold=True, size=10)
    h.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    h.cell(row=i, column=2, value=v).font = Font(name=ARIAL, size=9.5)
    h.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    h.row_dimensions[i].height = 62
    for j in (1, 2): h.cell(row=i, column=j).border = BORDER
h.column_dimensions["A"].width = 34; h.column_dimensions["B"].width = 108

out = os.path.join("outputs", "HYPOTHESIS_REGISTER.xlsx")
wb.save(out)
print("wrote", out, "| rows:", len(df), "| verdicts:", df["Verdict"].value_counts().to_dict())
